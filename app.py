"""
Sistema Operacional - Mika Transportes (versão Streamlit)
=============================================================
Aplicação interna (sem área pública) em um único arquivo Python:
  - Equipe operacional lança viagens/custos
  - Gerência acompanha um dashboard consolidado
  - Todos os dados ficam salvos localmente em cache.pkl (pickle)

Como rodar:
    pip install -r requirements.txt
    streamlit run app.py

Login de exemplo criado automaticamente na primeira execução:
    usuário: gerente     | senha: 123456   (perfil gerente)
    usuário: operacional | senha: 123456   (perfil operacional)
"""

import hashlib
import io
import json
import os
import pickle
from datetime import date, datetime

import pandas as pd
import plotly.graph_objects as go
import requests
import streamlit as st
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

# Carrega variáveis do arquivo .env quando rodando localmente.
# No Streamlit Cloud, as variáveis vêm de "Secrets" (Settings > Secrets do app),
# não deste arquivo.
load_dotenv()

# ----------------------------------------------------------------------
# Configuração e constantes
# ----------------------------------------------------------------------
CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache.pkl")
COMPROVANTES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Comprovantes")
os.makedirs(COMPROVANTES_DIR, exist_ok=True)

DB_URL = os.environ.get("DATABASE_URL")
if not DB_URL:
    st.error(
        "⚠️ Variável de ambiente DATABASE_URL não configurada.\n\n"
        "Local: crie um arquivo .env (veja .env.example) com DATABASE_URL=...\n"
        "Streamlit Cloud: configure em Settings > Secrets do app."
    )
    st.stop()

if DB_URL.startswith("postgres://"):
    DB_URL = DB_URL.replace("postgres://", "postgresql://", 1)

engine = create_engine(DB_URL)


def buscar_usuario_login(usuario_norm):
    """Busca um usuário no banco de dados pelo login (usuario). Retorna dict ou None."""
    with engine.connect() as conn:
        resultado = conn.execute(
            text("SELECT id, nome, usuario, senha_hash, perfil FROM usuarios WHERE usuario = :usuario"),
            {"usuario": usuario_norm},
        ).mappings().first()
    return dict(resultado) if resultado else None


def listar_usuarios():
    """Retorna todos os usuários cadastrados no banco (sem senha_hash)."""
    with engine.connect() as conn:
        resultado = conn.execute(
            text("SELECT id, nome, usuario, perfil FROM usuarios ORDER BY nome")
        ).mappings().all()
    return [dict(r) for r in resultado]


def buscar_usuario_por_id(usuario_id):
    """Busca um usuário pelo id (sem senha_hash), usado para edição."""
    with engine.connect() as conn:
        resultado = conn.execute(
            text("SELECT id, nome, usuario, perfil FROM usuarios WHERE id = :id"),
            {"id": usuario_id},
        ).mappings().first()
    return dict(resultado) if resultado else None


def usuario_login_existe(usuario_norm, excluir_id=None):
    """Verifica se já existe um usuário com esse login (usuario), ignorando excluir_id na edição."""
    with engine.connect() as conn:
        if excluir_id:
            resultado = conn.execute(
                text("SELECT id FROM usuarios WHERE usuario = :usuario AND id != :id"),
                {"usuario": usuario_norm, "id": excluir_id},
            ).first()
        else:
            resultado = conn.execute(
                text("SELECT id FROM usuarios WHERE usuario = :usuario"),
                {"usuario": usuario_norm},
            ).first()
    return resultado is not None


def criar_usuario_db(nome, usuario_norm, perfil, senha):
    with engine.connect() as conn:
        conn.execute(
            text("""
                INSERT INTO usuarios (nome, usuario, perfil, senha_hash)
                VALUES (:nome, :usuario, :perfil, :senha_hash)
            """),
            {"nome": nome, "usuario": usuario_norm, "perfil": perfil, "senha_hash": hash_senha(senha)},
        )
        conn.commit()


def atualizar_usuario_db(usuario_id, nome, usuario_norm, perfil, senha=None):
    with engine.connect() as conn:
        if senha:
            conn.execute(
                text("""
                    UPDATE usuarios SET nome=:nome, usuario=:usuario, perfil=:perfil, senha_hash=:senha_hash
                    WHERE id=:id
                """),
                {"nome": nome, "usuario": usuario_norm, "perfil": perfil,
                 "senha_hash": hash_senha(senha), "id": usuario_id},
            )
        else:
            conn.execute(
                text("""
                    UPDATE usuarios SET nome=:nome, usuario=:usuario, perfil=:perfil
                    WHERE id=:id
                """),
                {"nome": nome, "usuario": usuario_norm, "perfil": perfil, "id": usuario_id},
            )
        conn.commit()


def excluir_usuario_db(usuario_id):
    with engine.connect() as conn:
        conn.execute(text("DELETE FROM usuarios WHERE id = :id"), {"id": usuario_id})
        conn.commit()


def listar_usuarios():
    """Retorna todos os usuários cadastrados no banco (sem senha_hash)."""
    with engine.connect() as conn:
        resultado = conn.execute(
            text("SELECT id, nome, usuario, perfil FROM usuarios ORDER BY nome")
        ).mappings().all()
    return [dict(r) for r in resultado]


def buscar_usuario_por_id(usuario_id):
    """Busca um usuário pelo id (sem senha_hash), usado para edição."""
    with engine.connect() as conn:
        resultado = conn.execute(
            text("SELECT id, nome, usuario, perfil FROM usuarios WHERE id = :id"),
            {"id": usuario_id},
        ).mappings().first()
    return dict(resultado) if resultado else None


def usuario_login_existe(usuario_norm, excluir_id=None):
    """Verifica se já existe um usuário com esse login (usuario), ignorando excluir_id na edição."""
    with engine.connect() as conn:
        if excluir_id:
            resultado = conn.execute(
                text("SELECT id FROM usuarios WHERE usuario = :usuario AND id != :id"),
                {"usuario": usuario_norm, "id": excluir_id},
            ).first()
        else:
            resultado = conn.execute(
                text("SELECT id FROM usuarios WHERE usuario = :usuario"),
                {"usuario": usuario_norm},
            ).first()
    return resultado is not None


def criar_usuario_db(nome, usuario_norm, perfil, senha):
    with engine.connect() as conn:
        conn.execute(
            text("""
                INSERT INTO usuarios (nome, usuario, perfil, senha_hash)
                VALUES (:nome, :usuario, :perfil, :senha_hash)
            """),
            {"nome": nome, "usuario": usuario_norm, "perfil": perfil, "senha_hash": hash_senha(senha)},
        )
        conn.commit()


def atualizar_usuario_db(usuario_id, nome, usuario_norm, perfil, senha=None):
    with engine.connect() as conn:
        if senha:
            conn.execute(
                text("""
                    UPDATE usuarios SET nome=:nome, usuario=:usuario, perfil=:perfil, senha_hash=:senha_hash
                    WHERE id=:id
                """),
                {"nome": nome, "usuario": usuario_norm, "perfil": perfil,
                 "senha_hash": hash_senha(senha), "id": usuario_id},
            )
        else:
            conn.execute(
                text("""
                    UPDATE usuarios SET nome=:nome, usuario=:usuario, perfil=:perfil
                    WHERE id=:id
                """),
                {"nome": nome, "usuario": usuario_norm, "perfil": perfil, "id": usuario_id},
            )
        conn.commit()


def excluir_usuario_db(usuario_id):
    with engine.connect() as conn:
        conn.execute(text("DELETE FROM usuarios WHERE id = :id"), {"id": usuario_id})
        conn.commit()
CIDADES_CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cidades_brasil.json")


@st.cache_data(show_spinner=False)
def carregar_cidades_brasil():
    """Carrega a lista de cidades do Brasil no formato 'Cidade - UF'.
    Usa cache em disco (cidades_brasil.json) para não depender de internet
    toda vez que o app abrir. Retorna lista vazia se não conseguir (offline)."""
    if os.path.exists(CIDADES_CACHE_FILE):
        with open(CIDADES_CACHE_FILE, "r", encoding="utf-8") as f:
            cidades_salvas = json.load(f)
        if cidades_salvas:
            return cidades_salvas

    try:
        resposta = requests.get(
            "https://servicodados.ibge.gov.br/api/v1/localidades/municipios", timeout=15
        )
        resposta.raise_for_status()
        municipios = resposta.json()

        cidades_set = set()
        for m in municipios:
            try:
                nome = m["nome"]
                uf = m["microrregiao"]["mesorregiao"]["UF"]["sigla"]
            except (KeyError, TypeError):
                try:
                    uf = m["regiao-imediata"]["regiao-intermediaria"]["UF"]["sigla"]
                except (KeyError, TypeError):
                    continue
            cidades_set.add(f"{nome} - {uf}")

        cidades = sorted(cidades_set)

        if cidades:
            with open(CIDADES_CACHE_FILE, "w", encoding="utf-8") as f:
                json.dump(cidades, f, ensure_ascii=False)
        return cidades

    except Exception as erro:
        st.session_state["erro_cidades_brasil"] = str(erro)
        return []

GEOCODE_CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "geocode_cache.json")
DISTANCIAS_CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "distancias_cache.json")


def _carregar_json_cache(caminho):
    if os.path.exists(caminho):
        with open(caminho, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def _salvar_json_cache(caminho, dados_cache):
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados_cache, f, ensure_ascii=False)


def geocodificar_cidade(nome_cidade):
    """Converte 'Cidade - UF' em (latitude, longitude) usando Nominatim (OpenStreetMap).
    Usa cache em disco para evitar chamadas repetidas. Retorna None se não conseguir."""
    cache = _carregar_json_cache(GEOCODE_CACHE_FILE)
    if nome_cidade in cache:
        return tuple(cache[nome_cidade]) if cache[nome_cidade] else None

    try:
        resposta = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": f"{nome_cidade}, Brasil", "format": "json", "limit": 1},
            headers={"User-Agent": "MikaTransportesApp/1.0"},
            timeout=10,
        )
        resposta.raise_for_status()
        resultado = resposta.json()
        if not resultado:
            cache[nome_cidade] = None
            _salvar_json_cache(GEOCODE_CACHE_FILE, cache)
            return None

        coordenadas = (float(resultado[0]["lat"]), float(resultado[0]["lon"]))
        cache[nome_cidade] = coordenadas
        _salvar_json_cache(GEOCODE_CACHE_FILE, cache)
        return coordenadas
    except Exception:
        return None
def calcular_distancia_rodoviaria(origem, destino):
    """Calcula a distância rodoviária (KM) entre duas cidades 'Cidade - UF' via OSRM.
    Usa cache em disco. Retorna None se não conseguir calcular."""
    chave_cache = f"{origem}||{destino}"
    cache = _carregar_json_cache(DISTANCIAS_CACHE_FILE)
    if chave_cache in cache:
        return cache[chave_cache]

    coord_origem = geocodificar_cidade(origem)
    coord_destino = geocodificar_cidade(destino)
    if not coord_origem or not coord_destino:
        return None

    try:
        lat1, lon1 = coord_origem
        lat2, lon2 = coord_destino
        resposta = requests.get(
            f"http://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}",
            params={"overview": "false"},
            timeout=15,
        )
        resposta.raise_for_status()
        dados_rota = resposta.json()
        if dados_rota.get("code") != "Ok":
            return None

        distancia_km = round(dados_rota["routes"][0]["distance"] / 1000, 1)
        cache[chave_cache] = distancia_km
        _salvar_json_cache(DISTANCIAS_CACHE_FILE, cache)
        return distancia_km
    except Exception:
        return None
MESES_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
            "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

STATUS_OPCOES = {
    "em_transito": "Em Trânsito",
    "concluida": "Concluída",
    "atrasada": "Atrasada",
    "cancelada": "Cancelada",
}

COLUNAS_EXPORT = [
    "id", "data", "veiculo", "carreta", "motorista", "origem", "destino",
    "volume_tons", "faturamento_adiantamento", "faturamento_restante",
    "faturamento", "pedagio", "outros_custos", "custo_total", "status",
]
CABECALHO_EXPORT = [
    "ID", "Data", "Cavalo (Frota)", "Carreta", "Motorista", "Origem", "Destino",
    "Volume (T)", "Adiantamento (R$)", "Restante (R$)",
    "Faturamento Total (R$)", "Pedágio (R$)", "Outros Custos (R$)", "Custo Total (R$)", "Status",
]

COLUNAS_EXPORT_COMBUSTIVEL = [
    "id", "data", "veiculo", "motorista", "litros", "valor_pago",
    "valor_por_litro", "hodometro", "km_rodado", "cidade", "comprovante",
]
CABECALHO_EXPORT_COMBUSTIVEL = [
    "ID", "Data", "Veículo", "Motorista", "Litros", "Valor Pago (R$)",
    "R$/Litro", "Hodômetro (KM)", "KM Rodado", "Cidade", "Comprovante",
]

COR_RECEITA = "#2ecc71"
COR_DESPESA = "#e74c3c"
COR_LUCRO = "#f39c12"
COR_QTD = "#2d8eca"
COR_CARD_BG = "#1a2b47"
COR_CARD_BORDE = "#2d8eca"


def cartao_kpi(icone, titulo, valor, subtitulo=""):
    return f"""
    <div style="
        background: linear-gradient(135deg, {COR_CARD_BG}, #16213b);
        border: 1px solid {COR_CARD_BORDE}55;
        border-radius: 12px;
        padding: 18px 16px;
        text-align: center;
        height: 100%;
    ">
        <div style="font-size:1.5rem;">{icone}</div>
        <div style="color:#9fb3c8; font-size:0.75rem; text-transform:uppercase;
                    letter-spacing:0.04em; margin-top:6px;">{titulo}</div>
        <div style="color:#f2f2f2; font-size:1.55rem; font-weight:700; margin-top:2px;">{valor}</div>
        <div style="color:#7d92aa; font-size:0.72rem; margin-top:2px;">{subtitulo}</div>
    </div>
    """


def grafico_rdl_empilhado(receita_serie, despesa_serie, categoria_nome):
    """Gráfico de colunas empilhadas: Receita (verde) + Despesa (vermelho) + Lucro (laranja)."""
    categorias = sorted(set(receita_serie.index) | set(despesa_serie.index))
    receitas = [float(receita_serie.get(c, 0.0)) for c in categorias]
    despesas = [float(despesa_serie.get(c, 0.0)) for c in categorias]
    lucros = [r - d for r, d in zip(receitas, despesas)]

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=categorias, y=receitas, name="Receita", marker_color=COR_RECEITA,
        text=[f"R$ {v:,.0f}" for v in receitas], textposition="inside",
    ))
    fig.add_trace(go.Bar(
        x=categorias, y=despesas, name="Despesa", marker_color=COR_DESPESA,
        text=[f"R$ {v:,.0f}" for v in despesas], textposition="inside",
    ))
    fig.add_trace(go.Bar(
        x=categorias, y=lucros, name="Lucro", marker_color=COR_LUCRO,
        text=[f"R$ {v:,.0f}" for v in lucros], textposition="inside",
    ))
    fig.update_layout(
        barmode="stack",
        height=380,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#e6e6e6"),
        legend=dict(
            orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
            font=dict(color="#e6e6e6"),
        ),
        xaxis=dict(title=categoria_nome, gridcolor="#2a3f5f", tickangle=-20),
        yaxis=dict(title="R$", gridcolor="#2a3f5f"),
        margin=dict(t=30, b=40, l=40, r=20),
    )
    return fig


def grafico_quantidade(df_contagem, coluna_categoria, titulo_eixo):
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df_contagem[coluna_categoria], y=df_contagem["Quantidade"],
        marker_color=COR_QTD, text=df_contagem["Quantidade"], textposition="outside",
    ))
    fig.update_layout(
        height=320,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#e6e6e6"),
        legend=dict(font=dict(color="#e6e6e6")),
        xaxis=dict(title=titulo_eixo, gridcolor="#2a3f5f", tickangle=-20),
        yaxis=dict(title="Quantidade", gridcolor="#2a3f5f"),
        margin=dict(t=20, b=40, l=40, r=20),
    )
    return fig


COR_CUSTO_KM = "#8e44ad"


def grafico_custo_km(df_custo_km):
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df_custo_km["veiculo"], y=df_custo_km["custo_por_km"],
        marker_color=COR_CUSTO_KM,
        text=[f"R$ {v:,.2f}" for v in df_custo_km["custo_por_km"]], textposition="outside",
    ))
    fig.update_layout(
        height=350,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="#e6e6e6"),
        legend=dict(font=dict(color="#e6e6e6")),
        xaxis=dict(title="Caminhão", gridcolor="#2a3f5f", tickangle=-20),
        yaxis=dict(title="R$ / KM", gridcolor="#2a3f5f"),
        margin=dict(t=20, b=40, l=40, r=20),
    )
    return fig


# ----------------------------------------------------------------------
# Persistência dos dados (cache.pkl) - substitui o banco de dados
# ----------------------------------------------------------------------
def hash_senha(senha):
    """Hash simples da senha. Suficiente para uso interno; não é um cofre
    de segurança de nível bancário, mas evita salvar senha em texto puro."""
    return hashlib.sha256(senha.encode("utf-8")).hexdigest()


def checar_senha(senha, hash_salvo):
    return hash_senha(senha) == hash_salvo


def dados_iniciais():
    """Estrutura de exemplo usada na primeira execução (só login/usuários —
    as demais entidades vivem no banco de dados)."""
    return {
        "usuarios": [
            {"id": 1, "nome": "Gerente XYZ", "usuario": "gerente",
             "senha_hash": hash_senha("123456"), "perfil": "gerente"},
            {"id": 2, "nome": "Operacional XYZ", "usuario": "operacional",
             "senha_hash": hash_senha("123456"), "perfil": "operacional"},
        ],
    }


def carregar_dados():
    """Lê os usuários (login) do cache.pkl e as demais entidades direto do banco Postgres."""
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "rb") as f:
            dados = pickle.load(f)
        if migrar_dados(dados):
            salvar_dados(dados)
    else:
        dados = dados_iniciais()
        salvar_dados(dados)

    dados["motoristas"] = listar_motoristas()
    dados["veiculos"] = listar_veiculos()
    dados["carretas"] = listar_carretas()
    dados["viagens"] = listar_viagens()
    dados["abastecimentos"] = listar_abastecimentos()
    return dados


def migrar_dados(dados):
    """Ajusta caches antigos (.pkl) para a nova estrutura, sem apagar dados existentes.
    Agora cuida apenas de 'usuarios' (login) — as demais entidades vivem no banco.
    Retorna True se algo foi alterado (precisa salvar de novo)."""
    alterado = False

    for u in dados.get("usuarios", []):
        if "usuario" not in u:
            email_antigo = u.get("email", "")
            u["usuario"] = email_antigo.split("@")[0].strip().lower() if email_antigo else f"usuario{u['id']}"
            u.pop("email", None)
            alterado = True

    return alterado


def salvar_dados(dados):
    with open(CACHE_FILE, "wb") as f:
        pickle.dump(dados, f)


def proximo_id(lista):
    return max([item["id"] for item in lista], default=0) + 1


# ----------------------------------------------------------------------
# Persistência no banco de dados (Postgres via DATABASE_URL)
# ----------------------------------------------------------------------
@st.cache_data(ttl=300, show_spinner=False)
def listar_motoristas():
    with engine.connect() as conn:
        resultado = conn.execute(text("SELECT id, nome, cnh, telefone FROM motoristas ORDER BY nome"))
        return [dict(linha._mapping) for linha in resultado]


def criar_motorista(nome, cnh, telefone):
    with engine.begin() as conn:
        conn.execute(
            text("INSERT INTO motoristas (nome, cnh, telefone) VALUES (:nome, :cnh, :telefone)"),
            {"nome": nome, "cnh": cnh, "telefone": telefone},
        )
    listar_motoristas.clear()


def atualizar_motorista(motorista_id, nome, cnh, telefone):
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE motoristas SET nome=:nome, cnh=:cnh, telefone=:telefone WHERE id=:id"),
            {"nome": nome, "cnh": cnh, "telefone": telefone, "id": motorista_id},
        )
    listar_motoristas.clear()


def excluir_motorista(motorista_id):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM motoristas WHERE id=:id"), {"id": motorista_id})
    listar_motoristas.clear()


@st.cache_data(ttl=300, show_spinner=False)
def listar_veiculos():
    with engine.connect() as conn:
        resultado = conn.execute(text("SELECT id, placa, modelo, capacidade_kg FROM veiculos ORDER BY placa"))
        return [dict(linha._mapping) for linha in resultado]


def criar_veiculo(placa, modelo, capacidade_kg):
    with engine.begin() as conn:
        conn.execute(
            text("INSERT INTO veiculos (placa, modelo, capacidade_kg) VALUES (:placa, :modelo, :capacidade_kg)"),
            {"placa": placa, "modelo": modelo, "capacidade_kg": capacidade_kg},
        )
    listar_veiculos.clear()


def atualizar_veiculo(veiculo_id, placa, modelo, capacidade_kg):
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE veiculos SET placa=:placa, modelo=:modelo, capacidade_kg=:capacidade_kg WHERE id=:id"),
            {"placa": placa, "modelo": modelo, "capacidade_kg": capacidade_kg, "id": veiculo_id},
        )
    listar_veiculos.clear()


def excluir_veiculo(veiculo_id):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM veiculos WHERE id=:id"), {"id": veiculo_id})
    listar_veiculos.clear()


@st.cache_data(ttl=300, show_spinner=False)
def listar_carretas():
    with engine.connect() as conn:
        resultado = conn.execute(text("SELECT id, placa, modelo, capacidade_kg FROM carretas ORDER BY placa"))
        return [dict(linha._mapping) for linha in resultado]


def criar_carreta(placa, modelo, capacidade_kg):
    with engine.begin() as conn:
        conn.execute(
            text("INSERT INTO carretas (placa, modelo, capacidade_kg) VALUES (:placa, :modelo, :capacidade_kg)"),
            {"placa": placa, "modelo": modelo, "capacidade_kg": capacidade_kg},
        )
    listar_carretas.clear()


def atualizar_carreta(carreta_id, placa, modelo, capacidade_kg):
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE carretas SET placa=:placa, modelo=:modelo, capacidade_kg=:capacidade_kg WHERE id=:id"),
            {"placa": placa, "modelo": modelo, "capacidade_kg": capacidade_kg, "id": carreta_id},
        )
    listar_carretas.clear()


def excluir_carreta(carreta_id):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM carretas WHERE id=:id"), {"id": carreta_id})
    listar_carretas.clear()


@st.cache_data(ttl=120, show_spinner=False)
def listar_viagens():
    with engine.connect() as conn:
        resultado = conn.execute(text("SELECT * FROM viagens ORDER BY data DESC, id DESC"))
        return [dict(linha._mapping) for linha in resultado]


def criar_viagem(registro):
    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO viagens
                (data, veiculo_id, carreta_modelo, motorista_id, origem, destino, volume_tons,
                 faturamento_adiantamento, faturamento_restante, pedagio, outros_custos, status,
                 observacoes, criado_por_id)
                VALUES
                (:data, :veiculo_id, :carreta_modelo, :motorista_id, :origem, :destino, :volume_tons,
                 :faturamento_adiantamento, :faturamento_restante, :pedagio, :outros_custos, :status,
                 :observacoes, :criado_por_id)
            """),
            registro,
        )
    listar_viagens.clear()


def atualizar_viagem(viagem_id, registro):
    parametros = dict(registro, id=viagem_id)
    with engine.begin() as conn:
        conn.execute(
            text("""
                UPDATE viagens SET
                    data=:data, veiculo_id=:veiculo_id, carreta_modelo=:carreta_modelo,
                    motorista_id=:motorista_id, origem=:origem, destino=:destino,
                    volume_tons=:volume_tons, faturamento_adiantamento=:faturamento_adiantamento,
                    faturamento_restante=:faturamento_restante, pedagio=:pedagio,
                    outros_custos=:outros_custos, status=:status, observacoes=:observacoes
                WHERE id=:id
            """),
            parametros,
        )
    listar_viagens.clear()


def excluir_viagem(viagem_id):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM viagens WHERE id=:id"), {"id": viagem_id})
    listar_viagens.clear()


@st.cache_data(ttl=120, show_spinner=False)
def listar_abastecimentos():
    with engine.connect() as conn:
        resultado = conn.execute(text("""
            SELECT id, data, veiculo_id, motorista_id, litros, valor_pago, hodometro, cidade,
                   comprovante_nome, criado_por_id,
                   (comprovante_dados IS NOT NULL) AS tem_comprovante
            FROM abastecimentos
            ORDER BY data DESC, id DESC
        """))
        return [dict(linha._mapping) for linha in resultado]


def buscar_comprovante(abastecimento_id):
    with engine.connect() as conn:
        resultado = conn.execute(
            text("SELECT comprovante_nome, comprovante_dados FROM abastecimentos WHERE id=:id"),
            {"id": abastecimento_id},
        ).mappings().first()
    return dict(resultado) if resultado else None


def criar_abastecimento(registro, comprovante_nome=None, comprovante_bytes=None):
    parametros = dict(registro, comprovante_nome=comprovante_nome, comprovante_dados=comprovante_bytes)
    with engine.begin() as conn:
        conn.execute(
            text("""
                INSERT INTO abastecimentos
                (data, veiculo_id, motorista_id, litros, valor_pago, hodometro, cidade,
                 comprovante_nome, comprovante_dados, criado_por_id)
                VALUES
                (:data, :veiculo_id, :motorista_id, :litros, :valor_pago, :hodometro, :cidade,
                 :comprovante_nome, :comprovante_dados, :criado_por_id)
            """),
            parametros,
        )
    listar_abastecimentos.clear()


def atualizar_abastecimento(abastecimento_id, registro, comprovante_nome=None, comprovante_bytes=None):
    campos_extra = ""
    parametros = dict(registro, id=abastecimento_id)
    if comprovante_bytes is not None:
        campos_extra = ", comprovante_nome=:comprovante_nome, comprovante_dados=:comprovante_dados"
        parametros["comprovante_nome"] = comprovante_nome
        parametros["comprovante_dados"] = comprovante_bytes
    with engine.begin() as conn:
        conn.execute(
            text(f"""
                UPDATE abastecimentos SET
                    data=:data, veiculo_id=:veiculo_id, motorista_id=:motorista_id,
                    litros=:litros, valor_pago=:valor_pago, hodometro=:hodometro, cidade=:cidade
                    {campos_extra}
                WHERE id=:id
            """),
            parametros,
        )
    listar_abastecimentos.clear()


def excluir_abastecimento(abastecimento_id):
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM abastecimentos WHERE id=:id"), {"id": abastecimento_id})
    listar_abastecimentos.clear()

# ----------------------------------------------------------------------
# Funções auxiliares de consulta
# ----------------------------------------------------------------------
def buscar_por_id(lista, item_id):
    return next((item for item in lista if item["id"] == item_id), None)


def nome_veiculo(dados, veiculo_id):
    v = buscar_por_id(dados["veiculos"], veiculo_id)
    return v["placa"] if v else "-"


def nome_motorista(dados, motorista_id):
    m = buscar_por_id(dados["motoristas"], motorista_id)
    return m["nome"] if m else "-"


def nome_carreta(dados, carreta_id):
    c = buscar_por_id(dados["carretas"], carreta_id)
    return c["placa"] if c else "-"


def viagens_para_dataframe(dados):
    linhas = []
    for v in dados["viagens"]:
        adiantamento = float(v.get("faturamento_adiantamento", 0.0))
        restante = float(v.get("faturamento_restante", 0.0))
        pedagio = float(v.get("pedagio", 0.0))
        outros_custos = float(v.get("outros_custos", 0.0))
        linhas.append({
            "id": v["id"],
            "data": v["data"],
            "veiculo": nome_veiculo(dados, v["veiculo_id"]),
            "carreta": v.get("carreta_modelo") or "-",
            "motorista": nome_motorista(dados, v["motorista_id"]),
            "origem": v["origem"],
            "destino": v["destino"],
            "volume_tons": v["volume_tons"],
            "faturamento_adiantamento": adiantamento,
            "faturamento_restante": restante,
            "faturamento": adiantamento + restante,
            "pedagio": pedagio,
            "outros_custos": outros_custos,
            "custo_total": pedagio + outros_custos,
            "status": v["status"],
        })
    df = pd.DataFrame(linhas, columns=COLUNAS_EXPORT)
    if not df.empty:
        df["data"] = pd.to_datetime(df["data"])
    return df


# ----------------------------------------------------------------------
# Exportação (CSV / Excel / PDF)
# ----------------------------------------------------------------------
def gerar_csv(df):
    saida = io.StringIO()
    df_export = df.copy()
    df_export.columns = CABECALHO_EXPORT
    df_export["Data"] = df_export["Data"].dt.strftime("%d/%m/%Y")
    df_export["Status"] = df_export["Status"].replace(
        {v: r for v, r in [(k, STATUS_OPCOES[k]) for k in STATUS_OPCOES]}
    )
    df_export.to_csv(saida, index=False, sep=";")
    return ("\ufeff" + saida.getvalue()).encode("utf-8")


def gerar_xlsx(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Viagens"
    ws.append(CABECALHO_EXPORT)
    for celula in ws[1]:
        celula.font = Font(bold=True)
    for _, linha in df.iterrows():
        ws.append([
            int(linha["id"]), linha["data"].strftime("%d/%m/%Y"), linha["veiculo"],
            linha["carreta"], linha["motorista"], linha["origem"], linha["destino"],
            linha["volume_tons"],
            linha["faturamento_adiantamento"], linha["faturamento_restante"], linha["faturamento"],
            linha["pedagio"], linha["outros_custos"], linha["custo_total"],
            STATUS_OPCOES.get(linha["status"], linha["status"]),
        ])
    for coluna in ws.columns:
        largura = max((len(str(c.value)) for c in coluna if c.value is not None), default=10) + 2
        ws.column_dimensions[coluna[0].column_letter].width = min(largura, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_pdf(df):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    estilos = getSampleStyleSheet()

    linhas = [CABECALHO_EXPORT]
    for _, linha in df.iterrows():
        linhas.append([
            str(int(linha["id"])), linha["data"].strftime("%d/%m/%Y"), linha["veiculo"],
            linha["carreta"], linha["motorista"], linha["origem"], linha["destino"],
            str(linha["volume_tons"]),
            f"R$ {linha['faturamento_adiantamento']:.2f}", f"R$ {linha['faturamento_restante']:.2f}",
            f"R$ {linha['faturamento']:.2f}",
            f"R$ {linha['pedagio']:.2f}", f"R$ {linha['outros_custos']:.2f}", f"R$ {linha['custo_total']:.2f}",
            STATUS_OPCOES.get(linha["status"], linha["status"]),
        ])

    tabela = Table(linhas, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16213b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))

    elementos = [
        Paragraph("Relatório de Viagens — Mika Transportes", estilos["Title"]),
        Paragraph(f"Gerado em {date.today().strftime('%d/%m/%Y')}", estilos["Normal"]),
        tabela,
    ]
    doc.build(elementos)
    return buffer.getvalue()
def abastecimentos_para_dataframe(dados):
    # Calcula o KM rodado de cada abastecimento em relação ao abastecimento
    # anterior do MESMO caminhão (ordenado por data). O primeiro abastecimento
    # de cada caminhão fica sem KM rodado (não há referência anterior).
    abastecimentos_ordenados = sorted(
        dados["abastecimentos"], key=lambda a: (a["veiculo_id"], a["data"], a["id"])
    )
    km_rodado_por_id = {}
    hodometro_anterior_por_veiculo = {}
    for a in abastecimentos_ordenados:
        veiculo_id = a["veiculo_id"]
        hodometro_atual = float(a["hodometro"])
        if veiculo_id in hodometro_anterior_por_veiculo:
            km_rodado_por_id[a["id"]] = hodometro_atual - hodometro_anterior_por_veiculo[veiculo_id]
        else:
            km_rodado_por_id[a["id"]] = None
        hodometro_anterior_por_veiculo[veiculo_id] = hodometro_atual

    linhas = []
    for a in dados["abastecimentos"]:
        litros = float(a["litros"])
        valor_pago = float(a["valor_pago"])
        linhas.append({
            "id": a["id"],
            "data": a["data"],
            "veiculo": nome_veiculo(dados, a["veiculo_id"]),
            "motorista": nome_motorista(dados, a["motorista_id"]),
            "litros": litros,
            "valor_pago": valor_pago,
            "valor_por_litro": (valor_pago / litros) if litros else 0.0,
            "hodometro": a["hodometro"],
            "km_rodado": km_rodado_por_id.get(a["id"]),
            "cidade": a["cidade"],
            "comprovante": "Sim" if a.get("tem_comprovante") else "Não",
        })
    df = pd.DataFrame(linhas, columns=COLUNAS_EXPORT_COMBUSTIVEL)
    if not df.empty:
        df["data"] = pd.to_datetime(df["data"])
    return df


def gerar_csv_combustivel(df):
    saida = io.StringIO()
    df_export = df.copy()
    df_export.columns = CABECALHO_EXPORT_COMBUSTIVEL
    df_export["Data"] = df_export["Data"].dt.strftime("%d/%m/%Y")
    df_export.to_csv(saida, index=False, sep=";")
    return ("\ufeff" + saida.getvalue()).encode("utf-8")


def gerar_xlsx_combustivel(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Abastecimentos"
    ws.append(CABECALHO_EXPORT_COMBUSTIVEL)
    for celula in ws[1]:
        celula.font = Font(bold=True)
    for _, linha in df.iterrows():
        ws.append([
            int(linha["id"]), linha["data"].strftime("%d/%m/%Y"), linha["veiculo"], linha["motorista"],
            linha["litros"], linha["valor_pago"], round(linha["valor_por_litro"], 3),
            linha["hodometro"],
            round(linha["km_rodado"], 1) if pd.notna(linha["km_rodado"]) else "-",
            linha["cidade"], linha["comprovante"],
        ])
    for coluna in ws.columns:
        largura = max((len(str(c.value)) for c in coluna if c.value is not None), default=10) + 2
        ws.column_dimensions[coluna[0].column_letter].width = min(largura, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_pdf_combustivel(df):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    estilos = getSampleStyleSheet()

    linhas = [CABECALHO_EXPORT_COMBUSTIVEL]
    for _, linha in df.iterrows():
        linhas.append([
            str(int(linha["id"])), linha["data"].strftime("%d/%m/%Y"), linha["veiculo"], linha["motorista"],
            str(linha["litros"]), f"R$ {linha['valor_pago']:.2f}", f"R$ {linha['valor_por_litro']:.3f}",
            str(linha["hodometro"]),
            f"{linha['km_rodado']:.1f}" if pd.notna(linha["km_rodado"]) else "-",
            linha["cidade"], linha["comprovante"],
        ])

    tabela = Table(linhas, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16213b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))

    elementos = [
        Paragraph("Relatório de Abastecimentos — Mika Transportes", estilos["Title"]),
        Paragraph(f"Gerado em {date.today().strftime('%d/%m/%Y')}", estilos["Normal"]),
        tabela,
    ]
    doc.build(elementos)
    return buffer.getvalue()

# ----------------------------------------------------------------------
# Telas
# ----------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def _carregar_imagem_fundo_b64():
    import base64
    caminho_imagem = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "Imagens", "124-basculante.jpg"
    )
    with open(caminho_imagem, "rb") as f:
        return base64.b64encode(f.read()).decode()


def aplicar_fundo_com_imagem():
    """Aplica a imagem de fundo escurecida (login e menu inicial) e o estilo
    de botões em gradiente azul com fonte cinza claro."""
    imagem_b64 = _carregar_imagem_fundo_b64()

    st.markdown(
        f"""
        <style>
        [data-testid="stAppViewContainer"] {{
            background-image: linear-gradient(rgba(0,0,0,0.8), rgba(0,0,0,0.8)),
                               url("data:image/jpeg;base64,{imagem_b64}");
            background-size: cover;
            background-position: center;
            background-attachment: fixed;
        }}
        [data-testid="stHeader"] {{
            background: rgba(0,0,0,0);
        }}
        [data-testid="stAppViewContainer"] .block-container {{
            padding-top: 1rem;
        }}
        [data-testid="stForm"] label p {{
            color: #d9d9d9 !important;
        }}
        [data-testid="stForm"] {{
            background: rgba(255,255,255,0.06);
            border-radius: 10px;
            padding: 20px 25px;
        }}
        [data-testid="stForm"] button,
        div[data-testid="stButton"] button {{
            background: linear-gradient(135deg, #2d8eca, #1a4160);
            color: #d9d9d9 !important;
            border: none;
        }}
        [data-testid="stForm"] button:hover,
        div[data-testid="stButton"] button:hover {{
            background: linear-gradient(135deg, #1a4160, #2d8eca);
            color: #ffffff !important;
            border: none;
        }}
        [data-testid="stForm"] button p,
        div[data-testid="stButton"] button p {{
            color: inherit !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

def tela_login(dados):
    aplicar_fundo_com_imagem()

    st.markdown(
        "<h2 style='text-align:center; margin-top:60px; color:white;'>🚚 Mika Transportes</h2>"
        "<p style='text-align:center; color:#dddddd;'>Acesso restrito à equipe interna</p>",
        unsafe_allow_html=True,
    )
    col_esq, col_meio, col_dir = st.columns([1, 1, 1])
    with col_meio:
        with st.form("form_login"):
            usuario_input = st.text_input("Usuário")
            senha = st.text_input("Senha", type="password")
            entrar = st.form_submit_button("Entrar", use_container_width=True)

        if entrar:
            usuario_norm = usuario_input.strip().lower()
            usuario = buscar_usuario_login(usuario_norm)
            if usuario and checar_senha(senha, usuario["senha_hash"]):
                st.session_state.usuario = usuario
                st.rerun()
            else:
                st.error("Usuário ou senha inválidos.")
MENU_PRINCIPAL = [
    {"chave": "dashboard", "label": "📊 Dashboard", "somente_gerente": True},
    {"chave": "viagens", "label": "🚛 Viagens", "somente_gerente": False},
    {"chave": "combustivel", "label": "⛽ Combustível", "somente_gerente": False},
    {"chave": "motoristas", "label": "👤 Motoristas", "somente_gerente": True},
    {"chave": "frota", "label": "🚐 Frota", "somente_gerente": True},
    {"chave": "usuarios", "label": "🔑 Usuários", "somente_gerente": True},
]


def tela_menu_inicial(dados):
    aplicar_fundo_com_imagem()
    usuario = st.session_state.usuario

    st.markdown(
        f"<h1 style='text-align:center; margin-top:50px; color:white;'>🚚 Mika Transportes</h1>"
        f"<p style='text-align:center; color:#dddddd;'>Bem-vindo(a), {usuario['nome']}</p>",
        unsafe_allow_html=True,
    )
    st.write("")

    opcoes = [
        o for o in MENU_PRINCIPAL
        if not o["somente_gerente"] or usuario["perfil"] == "gerente"
    ]

    col_esq, col_meio, col_dir = st.columns([1, 1.2, 1])
    with col_meio:
        for opcao in opcoes:
            if st.button(opcao["label"], key=f"menu_{opcao['chave']}", use_container_width=True):
                st.session_state.pagina_atual = opcao["chave"]
                st.rerun()

        st.write("")
        if st.button("🚪 Sair", key="menu_sair", use_container_width=True):
            st.session_state.usuario = None
            st.session_state.pagina_atual = None
            st.rerun()

def formulario_viagem(dados, viagem=None):
    """Formulário de criação/edição de viagem. Se `viagem` for informado, edita."""
    motoristas = dados["motoristas"]
    veiculos = dados["veiculos"]
    carretas = dados["carretas"]

    modelos_disponiveis = sorted({c["modelo"].strip() for c in carretas if c["modelo"].strip()})
    if not motoristas or not veiculos or not modelos_disponiveis:
        st.warning(
            "Cadastre ao menos um motorista, um cavalo (frota) e uma carreta com modelo preenchido "
            "antes de lançar uma viagem."
        )
        return

    sufixo = viagem["id"] if viagem else "novo"

    chave_msg = f"msg_viagem_{sufixo}"
    if st.session_state.get(chave_msg):
        st.success(st.session_state.pop(chave_msg))

    cidades_brasil = carregar_cidades_brasil()

    with st.form(f"form_viagem_{sufixo}"):
        col1, col2 = st.columns(2)
        with col1:
            data_viagem = st.date_input(
                "Data da viagem *",
                value=pd.to_datetime(viagem["data"]).date() if viagem else date.today(),
            )
            if cidades_brasil:
                cidade_padrao_origem = "Três Lagoas - MS"
                indice_padrao_origem = (
                    cidades_brasil.index(cidade_padrao_origem)
                    if cidade_padrao_origem in cidades_brasil
                    else 0
                )
                origem = st.selectbox(
                    "Origem *", options=cidades_brasil,
                    index=(
                        cidades_brasil.index(viagem["origem"])
                        if viagem and viagem.get("origem") in cidades_brasil
                        else indice_padrao_origem
                    ),
                )
                destino = st.selectbox(
                    "Destino *", options=cidades_brasil,
                    index=(
                        cidades_brasil.index(viagem["destino"])
                        if viagem and viagem.get("destino") in cidades_brasil
                        else 0
                    ),
                )
            else:
                origem = st.text_input(
                    "Origem *", value=viagem["origem"] if viagem else "",
                    help="Lista automática indisponível (sem internet). Digite manualmente.",
                )
                destino = st.text_input(
                    "Destino *", value=viagem["destino"] if viagem else "",
                    help="Lista automática indisponível (sem internet). Digite manualmente.",
                )
            ids_veiculos = [v["id"] for v in veiculos]
            veiculo_id = st.selectbox(
                "Cavalo (Frota) *", options=ids_veiculos,
                format_func=lambda vid: next(f"{v['placa']} — {v['modelo']}" for v in veiculos if v["id"] == vid),
                index=(
                    ids_veiculos.index(viagem["veiculo_id"])
                    if viagem and viagem.get("veiculo_id") in ids_veiculos
                    else 0
                ),
            )
            carreta_modelo = st.selectbox(
                "Modelo da Carreta *", options=modelos_disponiveis,
                index=(
                    modelos_disponiveis.index(viagem["carreta_modelo"])
                    if viagem and viagem.get("carreta_modelo") in modelos_disponiveis
                    else 0
                ),
            )
            ids_motoristas = [m["id"] for m in motoristas]
            motorista_id = st.selectbox(
                "Motorista *", options=ids_motoristas,
                format_func=lambda mid: next(m["nome"] for m in motoristas if m["id"] == mid),
                index=(
                    ids_motoristas.index(viagem["motorista_id"])
                    if viagem and viagem.get("motorista_id") in ids_motoristas
                    else 0
                ),
            )
        with col2:
            status = st.selectbox(
                "Status *", options=list(STATUS_OPCOES.keys()),
                format_func=lambda s: STATUS_OPCOES[s],
                index=list(STATUS_OPCOES.keys()).index(viagem["status"]) if viagem else 0,
            )
            volume_tons = st.number_input(
                "Volume transportado (toneladas)", min_value=0.0, step=1.0,
                value=float(viagem["volume_tons"]) if viagem else 0.0,
            )
            faturamento_adiantamento = st.number_input(
                "Adiantamento de faturamento (R$) *", min_value=0.0, step=100.0,
                value=float(viagem.get("faturamento_adiantamento", 0.0)) if viagem else 0.0,
            )
            faturamento_restante = st.number_input(
                "Restante do faturamento (R$)", min_value=0.0, step=100.0,
                value=float(viagem.get("faturamento_restante", 0.0)) if viagem else 0.0,
                help="Opcional — pode lançar depois, quando receber o restante.",
            )
            pedagio = st.number_input(
                "Pedágio (R$)", min_value=0.0, step=10.0,
                value=float(viagem.get("pedagio", 0.0)) if viagem else 0.0,
                help="Opcional — pode lançar depois, quando souber o valor.",
            )
            outros_custos = st.number_input(
                "Outros custos (R$)", min_value=0.0, step=10.0,
                value=float(viagem.get("outros_custos", 0.0)) if viagem else 0.0,
                help="Opcional — pode lançar depois, quando souber o valor.",
            )
        observacoes = st.text_area("Observações", value=viagem.get("observacoes", "") if viagem else "")

        salvar = st.form_submit_button("💾 Salvar viagem", type="primary")

        if salvar:
            erros = []
            if not origem or not destino or not str(origem).strip() or not str(destino).strip():
                erros.append("Informe origem e destino.")
            if faturamento_adiantamento <= 0:
                erros.append("Informe o valor do adiantamento de faturamento.")
            if volume_tons <= 0:
                erros.append("Informe o volume transportado (toneladas).")

            if erros:
                for erro in erros:
                    st.error(erro)
                return

            registro = {
                "data": data_viagem.isoformat(),
                "veiculo_id": veiculo_id,
                "carreta_modelo": carreta_modelo,
                "motorista_id": motorista_id,
                "origem": origem.strip() if isinstance(origem, str) else origem,
                "destino": destino.strip() if isinstance(destino, str) else destino,
                "volume_tons": volume_tons,
                "faturamento_adiantamento": faturamento_adiantamento,
                "faturamento_restante": faturamento_restante,
                "pedagio": pedagio,
                "outros_custos": outros_custos,
                "status": status,
                "observacoes": observacoes.strip(),
            }

            if viagem is None:
                registro["criado_por_id"] = st.session_state.usuario["id"]
                criar_viagem(registro)
                st.session_state["msg_viagem_novo"] = "Viagem lançada com sucesso!"
            else:
                atualizar_viagem(viagem["id"], registro)
                st.session_state[f"msg_viagem_{viagem['id']}"] = "Viagem atualizada com sucesso!"

            st.session_state.pop("editando_viagem_id", None)
            st.rerun()
def formulario_abastecimento(dados, abastecimento=None):
    """Formulário de criação/edição de abastecimento. Se `abastecimento` for informado, edita."""
    motoristas = dados["motoristas"]
    veiculos = dados["veiculos"]

    if not motoristas or not veiculos:
        st.warning("Cadastre ao menos um motorista e um cavalo (frota) antes de lançar um abastecimento.")
        return

    if abastecimento and abastecimento.get("tem_comprovante"):
        comprovante_atual = buscar_comprovante(abastecimento["id"])
        if comprovante_atual and comprovante_atual.get("comprovante_dados"):
            st.image(comprovante_atual["comprovante_dados"], caption="Comprovante atual", width=220)

    with st.form(f"form_abastecimento_{abastecimento['id'] if abastecimento else 'novo'}"):
        col1, col2 = st.columns(2)
        with col1:
            data_abastecimento = st.date_input(
                "Data *",
                value=pd.to_datetime(abastecimento["data"]).date() if abastecimento else date.today(),
            )
            veiculo_id = st.selectbox(
                "Cavalo (Frota) *", options=[v["id"] for v in veiculos],
                format_func=lambda vid: next(f"{v['placa']} — {v['modelo']}" for v in veiculos if v["id"] == vid),
                index=[v["id"] for v in veiculos].index(abastecimento["veiculo_id"]) if abastecimento else 0,
            )
            motorista_id = st.selectbox(
                "Motorista *", options=[m["id"] for m in motoristas],
                format_func=lambda mid: next(m["nome"] for m in motoristas if m["id"] == mid),
                index=[m["id"] for m in motoristas].index(abastecimento["motorista_id"]) if abastecimento else 0,
            )
            cidades_brasil = carregar_cidades_brasil()
            if cidades_brasil:
                cidade_padrao_abastecimento = "Três Lagoas - MS"
                indice_padrao_abastecimento = (
                    cidades_brasil.index(cidade_padrao_abastecimento)
                    if cidade_padrao_abastecimento in cidades_brasil
                    else 0
                )
                cidade = st.selectbox(
                    "Cidade *", options=cidades_brasil,
                    index=(
                        cidades_brasil.index(abastecimento["cidade"])
                        if abastecimento and abastecimento.get("cidade") in cidades_brasil
                        else indice_padrao_abastecimento
                    ),
                )
            else:
                cidade = st.text_input(
                    "Cidade *", value=abastecimento["cidade"] if abastecimento else "",
                    help="Lista automática indisponível (sem internet). Digite manualmente.",
                )
        with col2:
            litros = st.number_input(
                "Litros *", min_value=0.0, step=1.0,
                value=float(abastecimento["litros"]) if abastecimento else 0.0,
            )
            valor_pago = st.number_input(
                "Valor pago (R$) *", min_value=0.0, step=10.0,
                value=float(abastecimento["valor_pago"]) if abastecimento else 0.0,
            )
            hodometro = st.number_input(
                "Hodômetro (KM) *", min_value=0.0, step=1.0,
                value=float(abastecimento["hodometro"]) if abastecimento else 0.0,
            )

        comprovante = st.file_uploader(
            "Foto do comprovante de abastecimento",
            type=["jpg", "jpeg", "png"],
            help="Opcional. Se não enviar uma nova, o comprovante atual (se houver) é mantido.",
        )

        salvar = st.form_submit_button("💾 Salvar abastecimento", type="primary")

        if salvar:
            erros = []
            if not cidade.strip():
                erros.append("Informe a cidade.")
            if litros <= 0:
                erros.append("Informe a quantidade de litros.")
            if valor_pago <= 0:
                erros.append("Informe o valor pago.")
            if hodometro <= 0:
                erros.append("Informe o hodômetro.")

            if erros:
                for erro in erros:
                    st.error(erro)
                return

            comprovante_nome = comprovante.name if comprovante is not None else None
            comprovante_bytes = comprovante.getvalue() if comprovante is not None else None

            registro = {
                "data": data_abastecimento.isoformat(),
                "veiculo_id": veiculo_id,
                "motorista_id": motorista_id,
                "litros": litros,
                "valor_pago": valor_pago,
                "hodometro": hodometro,
                "cidade": cidade.strip(),
            }

            if abastecimento is None:
                registro["criado_por_id"] = st.session_state.usuario["id"]
                criar_abastecimento(registro, comprovante_nome=comprovante_nome, comprovante_bytes=comprovante_bytes)
                st.success("Abastecimento lançado com sucesso!")
            else:
                atualizar_abastecimento(
                    abastecimento["id"], registro,
                    comprovante_nome=comprovante_nome, comprovante_bytes=comprovante_bytes,
                )
                st.success("Abastecimento atualizado com sucesso!")

            st.rerun()

def pagina_dashboard(dados):
    df = viagens_para_dataframe(dados)
    if df.empty:
        st.info("Nenhuma viagem lançada ainda. Lance a primeira em **Viagens**.")
        return

    df_abastecimentos = abastecimentos_para_dataframe(dados)

    st.subheader("Filtros")
    col_data_ini, col_data_fim, col_veiculo, col_motorista = st.columns(4)
    with col_data_ini:
        data_inicio = st.date_input("Data inicial", value=df["data"].min().date(), key="dash_data_inicio")
    with col_data_fim:
        data_fim = st.date_input("Data final", value=df["data"].max().date(), key="dash_data_fim")
    with col_veiculo:
        placas_disponiveis = sorted(df["veiculo"].dropna().unique().tolist())
        veiculo_sel = st.selectbox("Caminhão", options=["Todos"] + placas_disponiveis, key="dash_veiculo")
    with col_motorista:
        motoristas_disponiveis = sorted(df["motorista"].dropna().unique().tolist())
        motorista_sel = st.selectbox("Motorista", options=["Todos"] + motoristas_disponiveis, key="dash_motorista")

    df_filtrado = df[
        (df["data"].dt.date >= data_inicio) & (df["data"].dt.date <= data_fim)
    ].copy()
    if veiculo_sel != "Todos":
        df_filtrado = df_filtrado[df_filtrado["veiculo"] == veiculo_sel]
    if motorista_sel != "Todos":
        df_filtrado = df_filtrado[df_filtrado["motorista"] == motorista_sel]

    df_abast_filtrado = df_abastecimentos.copy()
    if not df_abast_filtrado.empty:
        df_abast_filtrado = df_abast_filtrado[
            (df_abast_filtrado["data"].dt.date >= data_inicio) & (df_abast_filtrado["data"].dt.date <= data_fim)
        ]
        if veiculo_sel != "Todos":
            df_abast_filtrado = df_abast_filtrado[df_abast_filtrado["veiculo"] == veiculo_sel]
        if motorista_sel != "Todos":
            df_abast_filtrado = df_abast_filtrado[df_abast_filtrado["motorista"] == motorista_sel]

    receita_total = df_filtrado["faturamento"].sum()
    despesa_viagens = df_filtrado["pedagio"].sum() + df_filtrado["outros_custos"].sum()
    despesa_combustivel = df_abast_filtrado["valor_pago"].sum() if not df_abast_filtrado.empty else 0.0
    despesa_total = despesa_viagens + despesa_combustivel
    lucro_total = receita_total - despesa_total
    qtd_lancamentos = len(df_filtrado)

    st.divider()
    card1, card2, card3, card4 = st.columns(4)
    with card1:
        st.markdown(cartao_kpi("💰", "Receita Total", f"R$ {receita_total:,.2f}"), unsafe_allow_html=True)
    with card2:
        st.markdown(cartao_kpi("💸", "Despesa Total", f"R$ {despesa_total:,.2f}"), unsafe_allow_html=True)
    with card3:
        st.markdown(cartao_kpi("📈", "Lucro (Receita - Despesa)", f"R$ {lucro_total:,.2f}"), unsafe_allow_html=True)
    with card4:
        st.markdown(cartao_kpi("🚛", "Viagens Lançadas", f"{qtd_lancamentos}"), unsafe_allow_html=True)

    st.divider()
    st.subheader("Receita, Despesa e Lucro por Caminhão e por Motorista")

    receita_veiculo = df_filtrado.groupby("veiculo")["faturamento"].sum()
    despesa_veiculo_viagem = (df_filtrado["pedagio"] + df_filtrado["outros_custos"]).groupby(df_filtrado["veiculo"]).sum()
    despesa_veiculo_combustivel = (
        df_abast_filtrado.groupby("veiculo")["valor_pago"].sum() if not df_abast_filtrado.empty else pd.Series(dtype=float)
    )
    despesa_veiculo = despesa_veiculo_viagem.add(despesa_veiculo_combustivel, fill_value=0.0)

    receita_motorista = df_filtrado.groupby("motorista")["faturamento"].sum()
    despesa_motorista_viagem = (df_filtrado["pedagio"] + df_filtrado["outros_custos"]).groupby(df_filtrado["motorista"]).sum()
    despesa_motorista_combustivel = (
        df_abast_filtrado.groupby("motorista")["valor_pago"].sum() if not df_abast_filtrado.empty else pd.Series(dtype=float)
    )
    despesa_motorista = despesa_motorista_viagem.add(despesa_motorista_combustivel, fill_value=0.0)

    col_g1, col_g2 = st.columns(2)
    with col_g1:
        st.caption("Por Caminhão")
        st.plotly_chart(grafico_rdl_empilhado(receita_veiculo, despesa_veiculo, "Caminhão"), use_container_width=True)
    with col_g2:
        st.caption("Por Motorista")
        st.plotly_chart(grafico_rdl_empilhado(receita_motorista, despesa_motorista, "Motorista"), use_container_width=True)

    st.divider()
    st.subheader("Receita, Despesa e Lucro por Mês")

    df_filtrado_mes = df_filtrado.copy()
    df_filtrado_mes["mes_ano"] = df_filtrado_mes["data"].dt.strftime("%Y-%m")
    receita_mes = df_filtrado_mes.groupby("mes_ano")["faturamento"].sum()
    despesa_mes_viagem = (df_filtrado_mes["pedagio"] + df_filtrado_mes["outros_custos"]).groupby(df_filtrado_mes["mes_ano"]).sum()
    if not df_abast_filtrado.empty:
        df_abast_mes = df_abast_filtrado.copy()
        df_abast_mes["mes_ano"] = df_abast_mes["data"].dt.strftime("%Y-%m")
        despesa_mes_combustivel = df_abast_mes.groupby("mes_ano")["valor_pago"].sum()
    else:
        despesa_mes_combustivel = pd.Series(dtype=float)
    despesa_mes = despesa_mes_viagem.add(despesa_mes_combustivel, fill_value=0.0)

    st.plotly_chart(grafico_rdl_empilhado(receita_mes, despesa_mes, "Mês"), use_container_width=True)

    st.divider()
    st.subheader("Quantidade de Viagens Lançadas")

    col_g3, col_g4 = st.columns(2)
    with col_g3:
        st.caption("Por Mês")
        qtd_mes = df_filtrado_mes.groupby("mes_ano").size().reset_index(name="Quantidade")
        st.plotly_chart(grafico_quantidade(qtd_mes, "mes_ano", "Mês"), use_container_width=True)
    with col_g4:
        st.caption("Por Caminhão")
        qtd_veiculo = df_filtrado.groupby("veiculo").size().reset_index(name="Quantidade")
        st.plotly_chart(grafico_quantidade(qtd_veiculo, "veiculo", "Caminhão"), use_container_width=True)

    st.divider()
    st.subheader("Custo por KM Rodado (por Caminhão)")

    km_veiculo = (
        df_abast_filtrado.groupby("veiculo")["km_rodado"].sum()
        if not df_abast_filtrado.empty else pd.Series(dtype=float)
    )
    linhas_custo_km = []
    for veiculo_nome in despesa_veiculo.index:
        km_total = km_veiculo.get(veiculo_nome, 0.0)
        if km_total and km_total > 0:
            custo_km = despesa_veiculo[veiculo_nome] / km_total
            linhas_custo_km.append({"veiculo": veiculo_nome, "custo_por_km": custo_km})

    if linhas_custo_km:
        df_custo_km = pd.DataFrame(linhas_custo_km).sort_values("custo_por_km", ascending=False)
        st.plotly_chart(grafico_custo_km(df_custo_km), use_container_width=True)
    else:
        st.info(
            "Sem quilometragem suficiente registrada (é necessário ao menos 2 abastecimentos "
            "por caminhão dentro do filtro selecionado) para calcular o custo por KM."
        )

    st.divider()
    st.subheader("Últimas viagens lançadas")
    ultimas = df.sort_values("data", ascending=False).head(6).copy()
    ultimas["data"] = ultimas["data"].dt.strftime("%d/%m/%Y")
    ultimas["status"] = ultimas["status"].map(STATUS_OPCOES)
    st.dataframe(
        ultimas.drop(columns=["id"]).rename(columns=dict(zip(COLUNAS_EXPORT[1:], CABECALHO_EXPORT[1:]))),
        use_container_width=True, hide_index=True,
    )

    st.divider()
    col_voltar, col_sair, _ = st.columns([1, 1, 4])
    with col_voltar:
        if st.button("🔙 Voltar ao menu inicial", key="voltar_dashboard", use_container_width=True):
            st.session_state.pagina_atual = None
            st.rerun()
    with col_sair:
        if st.button("🚪 Sair", key="sair_dashboard", use_container_width=True):
            st.session_state.usuario = None
            st.session_state.pagina_atual = None
            st.rerun()


def pagina_viagens(dados):
    st.title("Viagens")

    aba_nova, aba_lista = st.tabs(["➕ Nova viagem", "📋 Lançamentos"])

    with aba_nova:
        formulario_viagem(dados)

    with aba_lista:
        df = viagens_para_dataframe(dados)

        if df.empty:
            st.info("Nenhuma viagem cadastrada ainda.")
        else:
            tabela = df.sort_values("data", ascending=False).copy()
            tabela["data"] = tabela["data"].dt.strftime("%d/%m/%Y")
            tabela["status"] = tabela["status"].map(STATUS_OPCOES)
            st.dataframe(
                tabela.rename(columns=dict(zip(COLUNAS_EXPORT, CABECALHO_EXPORT))),
                use_container_width=True, hide_index=True,
            )

            st.divider()
            st.subheader("Editar ou excluir uma viagem")
            opcoes_id = [v["id"] for v in dados["viagens"]]
            if opcoes_id:
                viagem_id = st.selectbox(
                    "Selecione pelo ID", options=opcoes_id,
                    format_func=lambda vid: f"#{vid} — {buscar_por_id(dados['viagens'], vid)['origem']} → "
                                             f"{buscar_por_id(dados['viagens'], vid)['destino']}",
                )
                viagem_sel = buscar_por_id(dados["viagens"], viagem_id)

                with st.expander("✏️ Editar esta viagem"):
                    formulario_viagem(dados, viagem=viagem_sel)

                if st.session_state.usuario["perfil"] == "gerente":
                    if st.button("🗑️ Excluir viagem selecionada"):
                        excluir_viagem(viagem_id)
                        st.success("Viagem excluída.")
                        st.rerun()

            st.divider()
            st.subheader("Exportar")
            col_a, col_b, col_c, _ = st.columns([1, 1, 1, 3])
            with col_a:
                st.download_button("⬇ CSV", gerar_csv(df), "viagens.csv", "text/csv", use_container_width=True)
            with col_b:
                st.download_button(
                    "⬇ Excel", gerar_xlsx(df), "viagens.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with col_c:
                st.download_button("⬇ PDF", gerar_pdf(df), "viagens.pdf", "application/pdf", use_container_width=True)

    st.divider()
    col_voltar, col_sair, _ = st.columns([1, 1, 4])
    with col_voltar:
        if st.button("🔙 Voltar ao menu inicial", key="voltar_viagens", use_container_width=True):
            st.session_state.pagina_atual = None
            st.rerun()
    with col_sair:
        if st.button("🚪 Sair", key="sair_viagens", use_container_width=True):
            st.session_state.usuario = None
            st.session_state.pagina_atual = None
            st.rerun()


def pagina_motoristas(dados):
    st.title("Motoristas")
    col_form, col_lista = st.columns([1, 1.3])

    with col_form:
        editando_id = st.session_state.get("editando_motorista_id")
        motorista_edicao = buscar_por_id(dados["motoristas"], editando_id) if editando_id else None

        st.subheader("Editar motorista" if motorista_edicao else "Cadastrar motorista")
        with st.form("form_motorista"):
            nome = st.text_input("Nome *", value=motorista_edicao["nome"] if motorista_edicao else "")
            cnh = st.text_input("CNH", value=motorista_edicao["cnh"] if motorista_edicao else "")
            telefone = st.text_input("Telefone", value=motorista_edicao["telefone"] if motorista_edicao else "")
            salvar = st.form_submit_button("💾 Salvar", type="primary")
            if salvar:
                if not nome.strip():
                    st.error("Informe o nome do motorista.")
                else:
                    if motorista_edicao is None:
                        criar_motorista(nome.strip(), cnh.strip(), telefone.strip())
                        st.success("Motorista cadastrado.")
                    else:
                        atualizar_motorista(motorista_edicao["id"], nome.strip(), cnh.strip(), telefone.strip())
                        st.success("Motorista atualizado.")
                        st.session_state.pop("editando_motorista_id", None)

                    st.rerun()

        if motorista_edicao and st.button("Cancelar edição"):
            st.session_state.pop("editando_motorista_id", None)
            st.rerun()

    with col_lista:
        st.subheader("Motoristas cadastrados")
        if dados["motoristas"]:
            for m in dados["motoristas"]:
                c1, c2, c3, c4, c5 = st.columns([2, 1.3, 1.3, 0.5, 0.5])
                c1.write(m["nome"])
                c2.write(m["cnh"] or "-")
                c3.write(m["telefone"] or "-")
                if c4.button("✏️", key=f"edit_motorista_{m['id']}"):
                    st.session_state["editando_motorista_id"] = m["id"]
                    st.rerun()
                if st.session_state.usuario["perfil"] == "gerente":
                    if c5.button("🗑️", key=f"del_motorista_{m['id']}"):
                        excluir_motorista(m["id"])
                        st.success("Motorista excluído.")
                        st.rerun()
        else:
            st.info("Nenhum motorista cadastrado.")

    st.divider()
    col_voltar, col_sair, _ = st.columns([1, 1, 4])
    with col_voltar:
        if st.button("🔙 Voltar ao menu inicial", key="voltar_motoristas", use_container_width=True):
            st.session_state.pagina_atual = None
            st.rerun()
    with col_sair:
        if st.button("🚪 Sair", key="sair_motoristas", use_container_width=True):
            st.session_state.usuario = None
            st.session_state.pagina_atual = None
            st.rerun()


def pagina_veiculos(dados):
    st.title("Frota")

    aba_cavalos, aba_carretas = st.tabs(["🚚 Cavalos", "🚛 Carretas"])

    with aba_cavalos:
        col_form, col_lista = st.columns([1, 1.3])

        with col_form:
            editando_id = st.session_state.get("editando_veiculo_id")
            veiculo_edicao = buscar_por_id(dados["veiculos"], editando_id) if editando_id else None

            st.subheader("Editar cavalo" if veiculo_edicao else "Cadastrar cavalo (frota)")
            with st.form("form_veiculo"):
                placa = st.text_input(
                    "Placa *", placeholder="ABC1D23",
                    value=veiculo_edicao["placa"] if veiculo_edicao else "",
                )
                modelo = st.text_input("Modelo", value=veiculo_edicao["modelo"] if veiculo_edicao else "")
                capacidade_kg = st.number_input(
                    "Capacidade (KG)", min_value=0.0, step=100.0,
                    value=float(veiculo_edicao["capacidade_kg"]) if veiculo_edicao else 0.0,
                )
                salvar = st.form_submit_button("💾 Salvar", type="primary")
                if salvar:
                    if not placa.strip():
                        st.error("Informe a placa do cavalo.")
                    else:
                        if veiculo_edicao is None:
                            criar_veiculo(placa.strip().upper(), modelo.strip(), capacidade_kg)
                            st.success("Cavalo cadastrado.")
                        else:
                            atualizar_veiculo(veiculo_edicao["id"], placa.strip().upper(), modelo.strip(), capacidade_kg)
                            st.success("Cavalo atualizado.")
                            st.session_state.pop("editando_veiculo_id", None)

                        st.rerun()

            if veiculo_edicao and st.button("Cancelar edição", key="cancelar_edicao_veiculo"):
                st.session_state.pop("editando_veiculo_id", None)
                st.rerun()

        with col_lista:
            st.subheader("Cavalos cadastrados")
            if dados["veiculos"]:
                for v in dados["veiculos"]:
                    c1, c2, c3, c4, c5 = st.columns([1.5, 1.5, 1, 0.5, 0.5])
                    c1.write(v["placa"])
                    c2.write(v["modelo"] or "-")
                    c3.write(f"{v['capacidade_kg']:.0f} kg")
                    if c4.button("✏️", key=f"edit_veiculo_{v['id']}"):
                        st.session_state["editando_veiculo_id"] = v["id"]
                        st.rerun()
                    if st.session_state.usuario["perfil"] == "gerente":
                        if c5.button("🗑️", key=f"del_veiculo_{v['id']}"):
                            excluir_veiculo(v["id"])
                            st.success("Cavalo excluído.")
                            st.rerun()
            else:
                st.info("Nenhum cavalo cadastrado.")

    with aba_carretas:
        col_form, col_lista = st.columns([1, 1.3])

        with col_form:
            editando_id = st.session_state.get("editando_carreta_id")
            carreta_edicao = buscar_por_id(dados["carretas"], editando_id) if editando_id else None

            st.subheader("Editar carreta" if carreta_edicao else "Cadastrar carreta")
            with st.form("form_carreta"):
                placa = st.text_input(
                    "Placa da carreta *", placeholder="ABC1D23", key="placa_carreta",
                    value=carreta_edicao["placa"] if carreta_edicao else "",
                )
                modelo = st.text_input(
                    "Modelo *", placeholder="Ex: Basculante, Graneleira", key="modelo_carreta",
                    value=carreta_edicao["modelo"] if carreta_edicao else "",
                )
                capacidade_kg = st.number_input(
                    "Capacidade (KG)", min_value=0.0, step=100.0, key="capacidade_carreta",
                    value=float(carreta_edicao["capacidade_kg"]) if carreta_edicao else 0.0,
                )
                salvar = st.form_submit_button("💾 Salvar", type="primary")
                if salvar:
                    if not placa.strip():
                        st.error("Informe a placa da carreta.")
                    elif not modelo.strip():
                        st.error("Informe o modelo da carreta.")
                    else:
                        if carreta_edicao is None:
                            criar_carreta(placa.strip().upper(), modelo.strip(), capacidade_kg)
                            st.success("Carreta cadastrada.")
                        else:
                            atualizar_carreta(carreta_edicao["id"], placa.strip().upper(), modelo.strip(), capacidade_kg)
                            st.success("Carreta atualizada.")
                            st.session_state.pop("editando_carreta_id", None)

                        st.rerun()

            if carreta_edicao and st.button("Cancelar edição", key="cancelar_edicao_carreta"):
                st.session_state.pop("editando_carreta_id", None)
                st.rerun()

        with col_lista:
            st.subheader("Carretas cadastradas")
            if dados["carretas"]:
                for carreta in dados["carretas"]:
                    c1, c2, c3, c4, c5 = st.columns([1.5, 1.5, 1, 0.5, 0.5])
                    c1.write(carreta["placa"])
                    c2.write(carreta["modelo"] or "-")
                    c3.write(f"{carreta['capacidade_kg']:.0f} kg")
                    if c4.button("✏️", key=f"edit_carreta_{carreta['id']}"):
                        st.session_state["editando_carreta_id"] = carreta["id"]
                        st.rerun()
                    if st.session_state.usuario["perfil"] == "gerente":
                        if c5.button("🗑️", key=f"del_carreta_{carreta['id']}"):
                            excluir_carreta(carreta["id"])
                            st.success("Carreta excluída.")
                            st.rerun()
            else:
                st.info("Nenhuma carreta cadastrada.")

    st.divider()
    col_voltar, col_sair, _ = st.columns([1, 1, 4])
    with col_voltar:
        if st.button("🔙 Voltar ao menu inicial", key="voltar_frota", use_container_width=True):
            st.session_state.pagina_atual = None
            st.rerun()
    with col_sair:
        if st.button("🚪 Sair", key="sair_frota", use_container_width=True):
            st.session_state.usuario = None
            st.session_state.pagina_atual = None
            st.rerun()


def pagina_combustivel(dados):
    st.title("Combustível")

    aba_novo, aba_lista = st.tabs(["⛽ Novo abastecimento", "📋 Lançamentos"])

    with aba_novo:
        formulario_abastecimento(dados)

    with aba_lista:
        df = abastecimentos_para_dataframe(dados)

        if df.empty:
            st.info("Nenhum abastecimento cadastrado ainda.")
            return

        col_a, col_b, col_c, _ = st.columns([1, 1, 1, 3])
        with col_a:
            st.download_button(
                "⬇ CSV", gerar_csv_combustivel(df), "abastecimentos.csv", "text/csv",
                use_container_width=True,
            )
        with col_b:
            st.download_button(
                "⬇ Excel", gerar_xlsx_combustivel(df), "abastecimentos.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_c:
            st.download_button(
                "⬇ PDF", gerar_pdf_combustivel(df), "abastecimentos.pdf", "application/pdf",
                use_container_width=True,
            )

        st.divider()

        tabela = df.sort_values("data", ascending=False).copy()
        tabela["data"] = tabela["data"].dt.strftime("%d/%m/%Y")
        tabela["km_rodado"] = tabela["km_rodado"].apply(lambda v: f"{v:.0f}" if pd.notna(v) else "-")
        st.dataframe(
            tabela.rename(columns=dict(zip(COLUNAS_EXPORT_COMBUSTIVEL, CABECALHO_EXPORT_COMBUSTIVEL))),
            use_container_width=True, hide_index=True,
        )

        st.divider()
        st.subheader("Editar ou excluir um abastecimento")
        opcoes_id = [a["id"] for a in dados["abastecimentos"]]
        if opcoes_id:
            abastecimento_id = st.selectbox(
                "Selecione pelo ID", options=opcoes_id,
                format_func=lambda aid: f"#{aid} — {buscar_por_id(dados['abastecimentos'], aid)['cidade']} "
                                         f"({nome_veiculo(dados, buscar_por_id(dados['abastecimentos'], aid)['veiculo_id'])})",
            )
            abastecimento_sel = buscar_por_id(dados["abastecimentos"], abastecimento_id)

            with st.expander("✏️ Editar este abastecimento"):
                formulario_abastecimento(dados, abastecimento=abastecimento_sel)

            if st.session_state.usuario["perfil"] == "gerente":
                if st.button("🗑️ Excluir abastecimento selecionado"):
                    excluir_abastecimento(abastecimento_id)
                    st.success("Abastecimento excluído.")
                    st.rerun()

    st.divider()
    col_voltar, col_sair, _ = st.columns([1, 1, 4])
    with col_voltar:
        if st.button("🔙 Voltar ao menu inicial", key="voltar_combustivel", use_container_width=True):
            st.session_state.pagina_atual = None
            st.rerun()
    with col_sair:
        if st.button("🚪 Sair", key="sair_combustivel", use_container_width=True):
            st.session_state.usuario = None
            st.session_state.pagina_atual = None
            st.rerun()


def pagina_usuarios(dados):
    st.title("Usuários")

    if st.session_state.usuario["perfil"] != "gerente":
        st.error("Você não tem permissão para acessar esta página.")
        return

    col_form, col_lista = st.columns([1, 1.3])

    with col_form:
        editando_id = st.session_state.get("editando_usuario_id")
        usuario_edicao = buscar_usuario_por_id(editando_id) if editando_id else None

        st.subheader("Editar usuário" if usuario_edicao else "Cadastrar usuário")
        with st.form("form_usuario"):
            nome = st.text_input("Nome *", value=usuario_edicao["nome"] if usuario_edicao else "")
            login_usuario = st.text_input(
                "Nome de usuário *", value=usuario_edicao["usuario"] if usuario_edicao else ""
            )
            perfil = st.selectbox(
                "Perfil *", options=["operacional", "gerente"],
                index=["operacional", "gerente"].index(usuario_edicao["perfil"]) if usuario_edicao else 0,
            )
            senha = st.text_input(
                "Senha" + (" (deixe em branco para manter a atual)" if usuario_edicao else " *"),
                type="password",
            )
            salvar = st.form_submit_button("💾 Salvar", type="primary")

            if salvar:
                usuario_norm = login_usuario.strip().lower()
                if not nome.strip():
                    st.error("Informe o nome do usuário.")
                elif not usuario_norm:
                    st.error("Informe o nome de usuário.")
                elif usuario_login_existe(usuario_norm, excluir_id=usuario_edicao["id"] if usuario_edicao else None):
                    st.error("Já existe um usuário com este nome de usuário.")
                elif usuario_edicao is None and not senha:
                    st.error("Informe uma senha para o novo usuário.")
                else:
                    if usuario_edicao is None:
                        criar_usuario_db(nome.strip(), usuario_norm, perfil, senha)
                        st.success("Usuário cadastrado com sucesso!")
                    else:
                        atualizar_usuario_db(
                            usuario_edicao["id"], nome.strip(), usuario_norm, perfil,
                            senha=senha if senha else None,
                        )
                        st.success("Usuário atualizado com sucesso!")
                        st.session_state.pop("editando_usuario_id", None)

                    st.rerun()

        if usuario_edicao and st.button("Cancelar edição"):
            st.session_state.pop("editando_usuario_id", None)
            st.rerun()

    with col_lista:
        st.subheader("Usuários cadastrados")
        for u in listar_usuarios():
            c1, c2, c3, c4 = st.columns([2, 2, 1, 1])
            c1.write(u["nome"])
            c2.write(u["usuario"])
            c3.write(u["perfil"])
            with c4:
                if st.button("✏️", key=f"edit_{u['id']}"):
                    st.session_state["editando_usuario_id"] = u["id"]
                    st.rerun()
                if u["id"] != st.session_state.usuario["id"]:
                    if st.button("🗑️", key=f"del_{u['id']}"):
                        excluir_usuario_db(u["id"])
                        st.success("Usuário excluído.")
                        st.rerun()

    st.divider()
    col_voltar, col_sair, _ = st.columns([1, 1, 4])
    with col_voltar:
        if st.button("🔙 Voltar ao menu inicial", key="voltar_usuarios", use_container_width=True):
            st.session_state.pagina_atual = None
            st.rerun()
    with col_sair:
        if st.button("🚪 Sair", key="sair_usuarios", use_container_width=True):
            st.session_state.usuario = None
            st.session_state.pagina_atual = None
            st.rerun()


# ----------------------------------------------------------------------
# Navegação principal
# ----------------------------------------------------------------------
def barra_lateral():
    usuario = st.session_state.usuario
    st.sidebar.markdown("🚚 **MIKA TRANSPORTES**")
    st.sidebar.markdown(f"{usuario['nome']}  \n_{usuario['perfil']}_")
    st.sidebar.divider()

    if st.sidebar.button("🏠 Menu inicial", use_container_width=True):
        st.session_state.pagina_atual = None
        st.rerun()

    if st.sidebar.button("Sair", use_container_width=True):
        st.session_state.usuario = None
        st.session_state.pagina_atual = None
        st.rerun()


def main():
    st.set_page_config(page_title="Painel Operacional · Mika Transportes", layout="wide")

    if st.session_state.get("erro_cidades_brasil"):
        st.warning(f"⚠️ Não foi possível carregar cidades do Brasil: {st.session_state['erro_cidades_brasil']}")

    if "usuario" not in st.session_state:
        st.session_state.usuario = None
    if "pagina_atual" not in st.session_state:
        st.session_state.pagina_atual = None

    # Lê o cache.pkl a cada execução, garantindo dados sempre atualizados
    dados = carregar_dados()

    if st.session_state.usuario is None:
        tela_login(dados)
        return

    if st.session_state.pagina_atual is None:
        tela_menu_inicial(dados)
        return

    st.markdown(
        """
        <style>
        [data-testid="stAppViewContainer"] {
            background-image: none !important;
            background-color: #0f1c30 !important;
        }
        [data-testid="stAppViewContainer"] h1,
        [data-testid="stAppViewContainer"] h2,
        [data-testid="stAppViewContainer"] h3,
        [data-testid="stAppViewContainer"] p,
        [data-testid="stAppViewContainer"] label,
        [data-testid="stAppViewContainer"] .stMarkdown,
        [data-testid="stAppViewContainer"] .stCaption {
            color: #e6e6e6 !important;
        }
        [data-testid="stSidebar"] {
            background-color: #16213b !important;
        }
        [data-testid="stSidebar"] * {
            color: #e6e6e6 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    pagina = st.session_state.pagina_atual

    if pagina == "dashboard":
        pagina_dashboard(dados)
    elif pagina == "viagens":
        pagina_viagens(dados)
    elif pagina == "combustivel":
        pagina_combustivel(dados)
    elif pagina == "motoristas":
        pagina_motoristas(dados)
    elif pagina == "frota":
        pagina_veiculos(dados)
    elif pagina == "usuarios":
        pagina_usuarios(dados)


if __name__ == "__main__":
    main()
